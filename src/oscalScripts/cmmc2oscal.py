import openpyxl
# from oscal_pydantic.catalog import *
from oscalpy.catalog.catalog import *
from oscalpy.catalog.metadata import *
import datetime
import json

CONTROL_APPLICATIONS = (
    'rule',
    'mobile',
    'windows',
    'linux',
    'cloud'
)

def extract_control_from_cell(cell, ctrl_class=None) -> Control | None:
    if cell.value == None:
        return None
    v1ctrlid, v2ctrlid, ctrltitle, ctrldesc = cell.value.split('\n')[:4]
    ctrlrefs = cell.value.split('\n')[4:]
    
    descprop = Prop(
        name='description',
        value=ctrldesc,
        ns='cmmc',
        property_class='Control Metadata'
    )
    v1idprop = Prop(
        name='cmmc_v1_id',
        value=v1ctrlid,
        ns='cmmc',
        property_class='Control Metadata'
    )

    ctrl = Control(
        v2ctrlid,
        ctrltitle,
        ctrl_class,
    )
    ctrl.props = [descprop, v1idprop]

    for ref in ctrlrefs:
        prop = Prop(
            name='reference',
            value=ref,
            ns='cmmc',
            property_class='Control Metadata'
        )
        ctrl.props.append(prop)
    return ctrl

def main():

    party = Party(
        'organization',
        "DoD Cybersecurity Maturaity Model Certification (CMMC)",
        'CMMC',
    )
    party.add_link('https://dodcio.defense.gov/CMMC/', 'alternate', 'text/html', 'CMMC Homepage')

    l1_metadata = Metadata(
        title='CMMC Model V2 Level 1 Security Controls',
        version='2.0',
        oscal_version='1.1.1',
        last_modified=datetime.datetime.now().isoformat(),
        parties=[party]
    )

    l2_metadata = Metadata(
        title='CMMC Model V2 Level 2 Security Controls',
        version='2.0',
        oscal_version='1.1.1',
        parties=[party]
    )

    l3_metadata = Metadata(
        title='CMMC Model V2 Level 3 Security Controls',
        version='2.0',
        oscal_version='1.1.1',
        last_modified=datetime.datetime.now().isoformat(),
        parties=[party]
    )

    cmmc_v2_L1_catalog = Catalog("", "", "")
    cmmc_v2_L1_catalog.metadata = l1_metadata
    cmmc_v2_L1_catalog.groups = []
    cmmc_v2_L2_catalog = Catalog("", "", "")
    cmmc_v2_L2_catalog.metadata = l2_metadata
    cmmc_v2_L2_catalog.groups = []
    cmmc_v2_L3_catalog = Catalog("", "", "")
    cmmc_v2_L3_catalog.metadata = l3_metadata
    cmmc_v2_L3_catalog.groups = []

    
    # Load the workbook
    workbook = openpyxl.load_workbook('../ref/CMMCModel_V2_mapping.xlsx')

    # Iterate over the sheets
    for sheetname in workbook.sheetnames[1:-1]:
        sheet = workbook[sheetname]

        groupname = workbook[sheetname]['A1'].value.split('DOMAIN: ')[1]

        l1_group = Group(
            title=groupname,
            id=sheetname,
        )
        l1_group.controls = []
        l2_group = Group(
            id=sheetname,
            title=groupname,
        )
        l2_group.controls = []
        l3_group = Group(
            id=sheetname,
            title=groupname,
        )
        l3_group.controls = []
        # Iterate over the rows in the sheet
        for row in sheet.iter_rows(min_row=4, max_col=1):
            for cell in row:
                ctrl = extract_control_from_cell(cell, "L1")
                if ctrl != None:
                    l1_group.controls.append(ctrl)
                    l2_group.controls.append(ctrl)
                    l3_group.controls.append(ctrl)

        for row in sheet.iter_rows(min_row=4, min_col=2, max_col=2):
            for cell in row:
                ctrl = extract_control_from_cell(cell, "L2")
                
                if ctrl != None:
                    l2_group.controls.append(ctrl)
                    l3_group.controls.append(ctrl)
        
        for row in sheet.iter_rows(min_row=4, min_col=3, max_col=3):
            for cell in row:
                ctrl = extract_control_from_cell(cell, "L3")
                
                if ctrl != None:
                    l3_group.controls.append(ctrl)

        if len(l1_group.controls) != 0:
            cmmc_v2_L1_catalog.groups.append(l1_group)
        if len(l2_group.controls) != 0:
            cmmc_v2_L2_catalog.groups.append(l2_group)
        if len(l2_group.controls) != 0:
            cmmc_v2_L3_catalog.groups.append(l3_group)

    with open('cmmc_v2_L1.json', 'w') as f:
        json.dump(cmmc_v2_L1_catalog.as_dict(), f)
        f.close()
    with open('cmmc_v2_L2.json', 'w') as f:
        json.dump(cmmc_v2_L2_catalog.as_dict(), f)
        f.close()
    with open('cmmc_v2_L3.json', 'w') as f:
        json.dump(cmmc_v2_L3_catalog.as_dict(), f)
        f.close()

if __name__ == '__main__':
    main()
